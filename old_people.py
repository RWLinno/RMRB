"""
项目运行顺序:
    首先对每天爬取其有图片的详情页的url
    再依次对详情页url进行爬取
缺点:
    容易被封ip，时间间隔长
"""
import ddddocr
import asyncio
import re
import os
import time
import requests
import datetime
from lxml import etree
import random
from pyppeteer import launcher, launch
import openpyxl
from openpyxl import load_workbook

launcher.DEFAULT_ARGS.remove("--enable-automation")


def Make_Dir(now_year):
    """
    创建以年份为单位的文件夹与excel
    """
    if not os.path.exists(f'./output/{now_year}'):
        os.makedirs(f'./output/{now_year}')

        wb = openpyxl.Workbook()
        sh = wb.active
        sh.cell(column=1, row=1).value = '日期'
        sh.cell(column=2, row=1).value = '版次'
        sh.cell(column=3, row=1).value = '版面名称'
        sh.cell(column=4, row=1).value = '文章题目'
        sh.cell(column=5, row=1).value = '作者'
        sh.cell(column=6, row=1).value = '图片编号'
        sh.cell(column=7, row=1).value = '图片说明文字'
        sh.cell(column=8, row=1).value = '图片作者'
        wb.save(f'./output/{now_year}/info.xlsx')


def Save_Web_Txt(text, now_year, filename):
    """
    将网页返回值保存为HTML与TXT
    """
    filename = f'./output/{now_year}/' + filename + '.html'
    with open(filename, 'wb') as f:
        f.write(text.encode())
    filename = filename + '.txt'
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(text)


def Save_Pic(text, now_year, filename):
    """
    保存照片
    注意当一个Excel过大时保存会极其困难
    不过在此项目中一般不会出现这种情况
    """
    filename = f'./output/{now_year}/' + filename + '.png'
    with open(filename, 'wb') as f:
        f.write(text)


def Save_Excel(now_year, data_list):
    """
    保存Excel
    """
    wb = load_workbook(f'./output/{now_year}/info.xlsx')
    sh = wb.active
    max_ = sh.max_row
    for index_, single_data in enumerate(data_list):
        sh.cell(column=index_ + 1, row=max_ + 1).value = single_data
    wb.save(f'./output/{now_year}/info.xlsx')


async def Pass_Validate(url):
    """
    进行图片验证
    会弹出浏览器窗口 不可取消
    """
    browser = await launch(headless=False, dumpio=True, autoClose=False, args=['--start-maximized'])
    page = await browser.newPage()
    await page.setUserAgent(Random_User_Agent())
    await page.goto(url)
    while True:
        if await page.querySelector('#login_form > table > tbody > tr:nth-child(1) > td:nth-child(2) > img'):
            await page.screenshot(path='temp_screen.png', clip={'x': 550, 'y': 410, 'width': 100, 'height': 50})
            with open('temp_screen.png', 'rb') as f:
                img_bytes = f.read()
            await page.type('#validateCode', f'{OCR.classification(img_bytes)}',
                            {'delay': random.random()})
            await asyncio.sleep(random.random())
            await page.click('#login_form > table > tbody > tr:nth-child(2) > td:nth-child(2) > input',
                             options={'button': 'left', 'clickCount': 1})
        else:
            await asyncio.sleep(4)
            str_content = await page.content()
            break
        await asyncio.sleep(random.random())
    await browser.close()
    return str_content


def Get_Date_List(start_Date, end_Date):
    """
    获取日期列表方便遍历
    """

    start = datetime.datetime.strptime(start_Date, "%Y%m%d")
    end = datetime.datetime.strptime(end_Date, "%Y%m%d")

    temp_date_list = []
    for i in range((end - start).days + 1):
        temp_date_list.append(start + datetime.timedelta(days=1) * i)
    return temp_date_list


def Random_User_Agent():
    User_Agent_list = [
        'Mozilla/5.0(Macintosh;U;IntelMacOSX10_6_8;en-us)AppleWebKit/534.50(KHTML,likeGecko)Version/5.1Safari/534.50',
        'Opera/9.80(WindowsNT6.1;U;en)Presto/2.8.131Version/11.11',
        'Mozilla/5.0(Windows;U;WindowsNT6.1;en-us)AppleWebKit/534.50(KHTML,likeGecko)Version/5.1Safari/534.50',
        'Mozilla/5.0(Macintosh;IntelMacOSX10_7_0)AppleWebKit/535.11(KHTML,likeGecko)Chrome/17.0.963.56Safari/535.11',
        'Mozilla/4.0(compatible;MSIE7.0;WindowsNT5.1;Trident/4.0;SE2.XMetaSr1.0;SE2.XMetaSr1.0;.NETCLR2.0.50727;SE2.XMetaSr1.0)',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36',
        'Mozilla/4.0(compatible;MSIE7.0;WindowsNT5.1;360SE)',
        'Mozilla/4.0(compatible;MSIE7.0;WindowsNT5.1;Maxthon2.0)',
        'Opera/9.80(Macintosh;IntelMacOSX10.6.8;U;en)Presto/2.8.131Version/11.11',
    ]
    return random.choice(User_Agent_list)


def Get_Header():
    User_Agent_random = Random_User_Agent()
    headers = {
        "User-Agent": User_Agent_random,
        #'Cookie': 'JSESSIONID=67DA838A44DBE4386ED3BEFCA484E206; targetEncodinghttp://127001=2; validateCode=ls0hu5wesU9L6OXSOtzbwA%3D%3D',
        'Cookie' : 'wdcid=538d32400bd97bbd; targetEncodinghttp://127001=2; JSESSIONID=F19B58F0A3B7AE6F8A7137A889CC34F4',
        'Referer': 'http://data.people.com.cn/member/validateCode',
    }
    return headers


def Get_Pages_With_Pic(pages_html):
    """
    根据xpath中有无图片标识路径进行过滤
    不会点进详情页
    返回有图片的详情页的url列表
    """
    temp_html_list = []
    all_page_xpath = pages_html.xpath('//div[@class="title_list"]/ul/li/h3')
    for single_part in all_page_xpath:
        if single_part.xpath('./i'):
            temp_html_list.append(single_part.xpath('./a/@href')[0])
    return temp_html_list


def Get_Page_List(pages_url):
    """
    获得全部详情页的url列表并进行筛选
    返回删选后的url列表
    """
    html_list_day = []
    headers = Get_Header()
    try:
        print('正在获取第 1 版详情页信息')
        edition_1_response = requests.get(pages_url + '1', headers=headers).text
        edition_1_html = etree.HTML(edition_1_response)
        edition_num = int(edition_1_html.xpath('//div[@class="date"]/span[5]/text()')[0])
        html_list_day.extend(Get_Pages_With_Pic(edition_1_html))
        print('第 1 版详情页信息获取成功')

        for page_id in range(2, edition_num + 1):
            try:
                print(f'正在获取第 {page_id} 版详情页信息')
                temp_response = requests.get(pages_url + f'{page_id}', headers=headers).text
                temp_html = etree.HTML(temp_response)
                html_list_day.extend(Get_Pages_With_Pic(temp_html))
                print(f'第 {page_id} 版详情页信息获取成功')
            except Exception as e:
                print(f'第 {page_id} 版详情页信息获取失败')
                print(f'错误为: {e}')

    except Exception as e:
        print('第 1 版详情页信息获取失败')
        print(f'错误为: {e}')

    return html_list_day


def Handle_Single_Page(page_url):
    """
    对详情页中的信息进行爬取并进行保存
    """

    # 进行初次尝试查看是否要进行验证
    headers = Get_Header()
    response = requests.get('http://data.people.com.cn' + page_url, headers=headers)
    html_text = response.text
    html = etree.HTML(html_text)

    # 进行验证
    if html.xpath('//div[@class="bg_login_tb_div"]'):
        print('正在进行图片验证')
        html_text = asyncio.get_event_loop().run_until_complete(Pass_Validate('http://data.people.com.cn' + page_url))
        html = etree.HTML(html_text)

    # 日期
    publication_date = re.findall('rmrb/(\d+)/', page_url)[0]

    # 版次
    edition_num = html.xpath('//div[@class="sha_left"]/span[2]/text()')[0]
    if len(edition_num) >= 10:
        edition_num = '第' + edition_num + '版'
    else:
        edition_num = '第0' + edition_num + '版'

    # 版面名称
    edition_name = html.xpath('//div[@class="sha_left"]/span[3]/text()')
    if edition_name:
        edition_name = edition_num + ':' + edition_name[0]
    else:
        edition_name = edition_num + ':'

    # 标题
    title = '《' + html.xpath('//div[@class="title"]/text()')[0] + '》'
    author = html.xpath('//div[@class="author"]/text()')
    if author:
        author = author[0].replace('作者：', '').replace('【', '').replace('】', '')
    else:
        author = ''

    pics_info = html.xpath('//div[@align="center"]')
    temp_name = f'{publication_date}_{edition_num}_{title}'

    # 进行保存
    year_now = publication_date[:4]
    Make_Dir(year_now)
    Save_Web_Txt(html_text, year_now, temp_name)

    # 遍历图片进行保存
    for index_, pic_infor in enumerate(pics_info):

        pic_url = 'http://data.people.com.cn' + pic_infor.xpath('./img/@src')[0]
        words = ' '.join(pic_infor.xpath('.//td/text()')).replace('\u3000', ' ').replace('\t', '').strip()

        # 图片作者
        pic_author = re.findall('新华社记者 (.*)摄', words)
        if pic_author:
            pic_author = pic_author[0]
            content = words.split(f'新华社记者 {pic_author}摄')[0].strip()
            pic_author = pic_author.replace(' ', '')
        else:
            words = words.split(' ')
            if '摄' in words[-1]:
                pic_author = words[-1].replace('摄', '')
                content = ''.join(words[:-1]).strip()
            else:
                pic_author = ''
                content = ''.join(words).strip()

        # 图片编号
        pic_index = index_ + 1
        if pic_index < 10:
            pic_index = '0' + str(pic_index)

        # 保存图片
        pic_content = requests.get(pic_url, headers=Get_Header()).content
        Save_Pic(pic_content, year_now, temp_name + f'图片{pic_index}')

        # 保存Excel
        data_to_excel = [publication_date, edition_num, edition_name, title, author, pic_index, content, pic_author]
        Save_Excel(year_now, data_to_excel)

        print(f'URL: {page_url} 的页面已完成爬取')

def every_date(year):
    start_date = datetime.date(year, 4, 2)
    end_date = datetime.date(year, 12, 31)
    delta = datetime.timedelta(days=1)
    dates = []
    while start_date <= end_date:
        dates.append(start_date.strftime('%Y-%m-%d'))
        start_date += delta
    return dates

def main():
    """
    主函数:程序入口
    """
    # 输入起止日期，爬取之间的新闻 类似1990-1-1
    # beginDate = input('请输入开始日期:')
    # endDate = input('请输入结束日期:')
    # date_list = Get_Date_List(beginDate, endDate)

    # 对某一年全部日期爬取
    # date_list = every_date(year)  
    
    # 指定日期
    date_list = ['19900101','19801213']
    for single_day in date_list:
        year = str(single_day[:4])
        month = str(single_day[4:6])
        day = str(single_day[6:8])
        #year = str(single_day.year)
        #month = str(single_day.month) if single_day.month >= 10 else '0' + str(single_day.month)
        #day = str(single_day.day) if single_day.day >= 10 else '0' + str(single_day.day)
        single_day_url = 'http://data.people.com.cn/rmrb/' + year + month + day + '/'
        print(single_day_url)
        for single_link in Get_Page_List(single_day_url):  # 每一个版面的链接
            error_time = 0
            while error_time < 3:
                try:
                    Handle_Single_Page(single_link)
                    time.sleep(10)
                    break
                except Exception as e:
                    error_time += 1
                    print(f'页面 {single_link} 当前第 {error_time} 次爬取失败')
                    print(f'错误为: {e}')
            if error_time == 3:
                print('!!!!!!请注意IP封禁!!!!!!')
                return
        print(f'爬取完成 {year} 年 {month} 月 {day} 日信息')


if __name__ == '__main__':
    # 文字识别器
    OCR = ddddocr.DdddOcr()
    main()
